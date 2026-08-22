#!/usr/bin/env python3
"""
cargar_padron_julio.py — Carga el PADRON de contadores/estudios (corte 20-jul-2026)
a contadores_padron. Formato SUNAT: cabecera en filas 1-7, datos desde la fila 8.

Mismo layout de columnas que los CONTADORES-PERU-PERS anteriores:
  Naturales (CONTADORES-20JULIO2026.xlsx, 13 cols):
    RUC(0), Razon(1), TipoContrib(2), Estado(3), CIIU(4), DescCIIU(5), Ubigeo(6),
    Distrito(7), Provincia(8), Departamento(9), Dependencia(10), tributo(11), NombreComercial(12)
  Juridicas / estudios (ESTUDIOS-CONTABLES-20JULIO2026.xlsx, 24 cols):
    RUC(0), Razon(1), FInsc(2), FInicio(3), TipoContrib(4), Estado(5), CIIU(6), Desc(7),
    TipoVia(8), NombreVia(9), Numero(10), Interior(11), TipoZona(12), NombreZona(13),
    kilom(14), manza(15), depar(16), lote(17), Ubigeo(18), Distrito(19), Provincia(20),
    Departamento(21), Tributo(22), NombreComercial(23)  -> SI traen direccion.

Reglas:
  - tipo derivado del RUC: 10/15/17 = natural, 20 = juridica.
  - DISTRITO/PROV/DEP de display salen del UBIGEO (cat de distritos.json), NO del
    texto (la ñ del archivo viene rota). Fallback al texto solo si el ubigeo no
    esta en el catalogo.
  - Estudios juridicos: se arma direccion_texto (columna que se agrega si falta).
  - Dedup por RUC: ON CONFLICT (ruc) DO NOTHING. NO toca filas existentes ni los
    campos perfil_* (solo inserta RUCs nuevos).

Uso:
  python cargar_padron_julio.py --naturales <ruta> --juridicas <ruta> --dry-run
  python cargar_padron_julio.py --naturales <ruta> --juridicas <ruta>
"""
from __future__ import annotations

import argparse
import asyncio
import json
import os
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("Falta openpyxl. Instala:  pip install openpyxl")

BASE = Path(__file__).resolve().parent
DISTRITOS_JSON = BASE / "app" / "static" / "distritos.json"
FILA_DATOS = 8
LOTE = 1000

COLS_NATURAL = dict(ruc=0, razon_social=1, tipo_contrib=2, estado=3,
                    ubigeo=6, distrito=7, provincia=8, departamento=9, nombre_comercial=12)
COLS_JURIDICA = dict(ruc=0, razon_social=1, tipo_contrib=4, estado=5,
                     tipo_via=8, nombre_via=9, numero=10, interior=11,
                     tipo_zona=12, nombre_zona=13, km=14, mz=15, num_depar=16, num_lote=17,
                     ubigeo=18, distrito=19, provincia=20, departamento=21, nombre_comercial=23)


def txt(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s and s != "-" else None


def digitos(v):
    return "".join(c for c in str(v) if c.isdigit()) if v is not None else ""


def cod2(v):
    d = digitos(v)
    return d.zfill(2)[-2:] if d else None


def cargar_dotenv():
    env = BASE / ".env"
    if not env.exists():
        return
    for l in env.read_text(encoding="utf-8").splitlines():
        l = l.strip()
        if l and not l.startswith("#") and "=" in l:
            k, v = l.split("=", 1)
            os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))


def cat_ubigeo():
    d = json.loads(DISTRITOS_JSON.read_text(encoding="utf-8"))
    return {x["u"]: (x["d"], x["p"], x["dep"]) for x in d}


def _cap(s):
    return (s or "").strip()


def armar_direccion(g, via_cat, zona_cat):
    """Direccion legible del estudio con via/zona decodificadas (cat_via/cat_zona)."""
    via_den = via_cat.get(cod2(g("tipo_via")), "")
    zona_den = zona_cat.get(cod2(g("tipo_zona")), "")
    partes = []
    l1 = " ".join(p for p in [via_den, txt(g("nombre_via"))] if p).strip()
    num = txt(g("numero"))
    if num and num.upper() not in ("S/N", "SN", "0"):
        l1 = f"{l1} Nro. {num}".strip()
    if l1:
        partes.append(l1)
    extras = []
    for etq, k in (("Int.", "interior"), ("Dpto.", "num_depar"),
                   ("Km.", "km"), ("Mz.", "mz"), ("Lt.", "num_lote")):
        val = txt(g(k))
        if val:
            extras.append(f"{etq} {val}")
    if extras:
        partes.append(" ".join(extras))
    if zona_den and zona_den != "Otros":
        z = f"{zona_den} {txt(g('nombre_zona')) or ''}".strip()
        if z:
            partes.append(z)
    dire = ", ".join(p for p in partes if p and p.strip())
    dire = re.sub(r"\s{2,}", " ", dire).strip(" ,-")
    return dire or None


def leer(ruta, mapa, es_juridica, catub, via_cat, zona_cat):
    wb = load_workbook(ruta, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    filas = []
    sin_ubigeo = 0
    for row in ws.iter_rows(min_row=FILA_DATOS, values_only=True):
        if row is None or all(c is None for c in row):
            continue

        def g(k):
            i = mapa.get(k)
            return row[i] if (i is not None and i < len(row)) else None

        ruc = digitos(g("ruc"))
        if len(ruc) != 11:
            continue
        ubigeo = digitos(g("ubigeo")).zfill(6)[-6:] if g("ubigeo") else None
        cat = catub.get(ubigeo) if ubigeo else None
        if ubigeo and not cat:
            sin_ubigeo += 1
        distrito = cat[0] if cat else txt(g("distrito"))
        provincia = cat[1] if cat else txt(g("provincia"))
        departamento = cat[2] if cat else txt(g("departamento"))
        tipo = "juridica" if ruc[:2] == "20" else "natural"
        direccion = armar_direccion(g, via_cat, zona_cat) if es_juridica else None
        filas.append({
            "ruc": ruc, "razon_social": txt(g("razon_social")), "tipo": tipo,
            "estado": txt(g("estado")), "ubigeo": ubigeo,
            "distrito": distrito, "provincia": provincia, "departamento": departamento,
            "nombre_comercial": txt(g("nombre_comercial")), "direccion_texto": direccion,
        })
    wb.close()
    return filas, sin_ubigeo


INSERT_COLS = ["ruc", "razon_social", "tipo", "estado", "ubigeo",
               "distrito", "provincia", "departamento", "nombre_comercial", "direccion_texto"]


async def cat_dicts(conn):
    via = {r["codigo"]: r["denominacion"] for r in await conn.fetch("SELECT codigo,denominacion FROM cat_via")}
    zona = {r["codigo"]: r["denominacion"] for r in await conn.fetch("SELECT codigo,denominacion FROM cat_zona")}
    return via, zona


async def analizar(database_url, rutas):
    """Lee catalogos de la BD, mapea, y reporta SIN insertar (incluye solape)."""
    import asyncpg
    conn = await asyncpg.connect(database_url)
    try:
        via_cat, zona_cat = await cat_dicts(conn)
        catub = cat_ubigeo()
        nat, snat = leer(rutas["naturales"], COLS_NATURAL, False, catub, via_cat, zona_cat)
        jur, sjur = leer(rutas["juridicas"], COLS_JURIDICA, True, catub, via_cat, zona_cat)
        filas = nat + jur
        rucs = [f["ruc"] for f in filas]
        ya = set(await conn.fetchval(
            "SELECT array_agg(ruc) FROM contadores_padron WHERE ruc = ANY($1::text[])", rucs) or [])
        total_actual = await conn.fetchval("SELECT COUNT(*) FROM contadores_padron")
        iq_actual = await conn.fetchval("SELECT COUNT(*) FROM contadores_padron WHERE left(ubigeo,4)='1601'")
        iq_files = sum(1 for f in filas if (f["ubigeo"] or "")[:4] == "1601")
        iq_nuevos = sum(1 for f in filas if (f["ubigeo"] or "")[:4] == "1601" and f["ruc"] not in ya)
        return {"nat": nat, "jur": jur, "snat": snat, "sjur": sjur,
                "n_ya": len(ya), "n_nuevos": len(filas) - len(ya),
                "total_actual": total_actual, "iq_actual": iq_actual,
                "iq_files": iq_files, "iq_nuevos": iq_nuevos}
    finally:
        await conn.close()


async def cargar(database_url, filas):
    import asyncpg
    conn = await asyncpg.connect(database_url)
    try:
        await conn.execute("ALTER TABLE contadores_padron ADD COLUMN IF NOT EXISTS direccion_texto text")
        antes = await conn.fetchval("SELECT COUNT(*) FROM contadores_padron")
        ph = ", ".join(f"${i}" for i in range(1, len(INSERT_COLS) + 1))
        sql = (f"INSERT INTO contadores_padron ({', '.join(INSERT_COLS)}) "
               f"VALUES ({ph}) ON CONFLICT (ruc) DO NOTHING")
        for i in range(0, len(filas), LOTE):
            lote = filas[i:i + LOTE]
            await conn.executemany(sql, [[f[c] for c in INSERT_COLS] for f in lote])
        despues = await conn.fetchval("SELECT COUNT(*) FROM contadores_padron")
        return antes, despues
    finally:
        await conn.close()


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--naturales", required=True)
    ap.add_argument("--juridicas", required=True)
    ap.add_argument("--database-url", default=None)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()
    for r in (args.naturales, args.juridicas):
        if not Path(r).exists():
            sys.exit(f"No existe: {r}")

    cargar_dotenv()
    database_url = args.database_url or os.getenv("DATABASE_URL", "").strip()
    if not database_url:
        sys.exit("Falta DATABASE_URL.")
    rutas = {"naturales": args.naturales, "juridicas": args.juridicas}

    a = asyncio.run(analizar(database_url, rutas))
    filas = a["nat"] + a["jur"]
    print(f"Naturales leidas: {len(a['nat'])}  (ubigeo fuera de catalogo: {a['snat']})")
    print(f"Juridicas leidas: {len(a['jur'])}  (ubigeo fuera de catalogo: {a['sjur']})")
    print(f"Total en archivos: {len(filas)}")
    print(f"  Ya existen en el padron (no se insertan): {a['n_ya']}")
    print(f"  Nuevos a insertar:                        {a['n_nuevos']}")
    print(f"  Padron actual: {a['total_actual']}  ->  proyectado: {a['total_actual'] + a['n_nuevos']}")
    print(f"  IQUITOS (1601) en archivos: {a['iq_files']}  | ya en padron 1601: {a['iq_actual']}  "
          f"| nuevos 1601: {a['iq_nuevos']}")
    print("  Muestra NATURAL:", a["nat"][0])
    print("  Muestra JURIDICA:", a["jur"][0])

    if args.dry_run:
        print("DRY-RUN: no se inserto nada.")
        return

    antes, despues = asyncio.run(cargar(database_url, filas))
    print(f"\nListo. contadores_padron: {antes} -> {despues} (+{despues-antes} insertados).")


if __name__ == "__main__":
    main()
