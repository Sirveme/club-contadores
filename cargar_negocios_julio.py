#!/usr/bin/env python3
"""
cargar_negocios_julio.py — Cargador de NUEVOS NEGOCIOS de JULIO 2026.

El formato de julio es DISTINTO al de mayo/junio (ver cargar_negocios.py):
  - NO trae columna ubigeo -> se DERIVA por nombre contra app/static/distritos.json.
  - NO trae codigo de tributo -> el regimen de JURIDICAS se deriva de regimen_desc
    (texto) a tributo_codigo (4 dig), para que el reporte lo agrupe igual que
    mayo/junio. Las NATURALES no traen regimen -> queda NULL (honesto).
  - Naturales: distrito en columnas Departamento/Provincia/Distrito.
  - Juridicas: distrito EMBEBIDO al final de 'domicilio' (3 ultimos segmentos =
    DISTRITO, PROVINCIA, DEPARTAMENTO). El domicilio completo -> direccion_texto.

Resolucion de ubigeo (nombre -> ubigeo de 6 digitos), en cascada:
  1) triple exacto (departamento, provincia, distrito) normalizado
  2) (provincia, distrito) si es unico en el pais
  3) distrito solo si es unico en el pais
  Ademas: quita parentesis ("PUEBLO LIBRE (MAGDALENA VIEJA)") probando variantes,
  y aplica un mapa corto de ALIAS de abreviaturas de SUNAT.
  Lo que no resuelve -> ubigeo NULL (la fila se carga igual y se registra aparte).

Idempotente: ON CONFLICT (ruc) DO NOTHING. No toca mayo/junio.

Uso:
  # Muestra/diagnostico SIN insertar (incluye desglose de los 4 distritos de Iquitos):
  python cargar_negocios_julio.py --tipo juridica --archivo <ruta> --dry-run
  python cargar_negocios_julio.py --tipo natural  --archivo <ruta> --dry-run
  # Carga real por lotes:
  python cargar_negocios_julio.py --tipo juridica --archivo <ruta>
  python cargar_negocios_julio.py --tipo natural  --archivo <ruta>
"""
from __future__ import annotations

import argparse
import asyncio
import datetime as dt
import json
import os
import re
import sys
import unicodedata
from collections import Counter
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("Falta openpyxl. Instala:  pip install openpyxl")

BASE = Path(__file__).resolve().parent
DISTRITOS_JSON = BASE / "app" / "static" / "distritos.json"
PERIODO = "2026-07"
LOTE = 1000


# --- .env minimalista --------------------------------------------------------
def cargar_dotenv() -> None:
    env = BASE / ".env"
    if not env.exists():
        return
    for linea in env.read_text(encoding="utf-8").splitlines():
        linea = linea.strip()
        if linea and not linea.startswith("#") and "=" in linea:
            k, v = linea.split("=", 1)
            os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))


# --- Helpers de valor --------------------------------------------------------
def texto(v):
    if v is None:
        return None
    try:
        if isinstance(v, float) and v.is_integer():
            v = int(v)
        s = str(v)
        s = s.encode("utf-8", "replace").decode("utf-8", "replace")
        s = "".join(c for c in s if c in "\n\t" or ord(c) >= 32).strip()
        # SUNAT usa '-' como "vacio" en NombreComercial.
        return None if s in ("", "-") else s
    except Exception:
        return None


def mayus(v):
    s = texto(v)
    return s.upper() if s else None


def minus(v):
    s = texto(v)
    return s.lower() if s else None


def ruc_norm(v):
    d = re.sub(r"\D", "", str(v)) if v is not None else ""
    if not d:
        return None
    return d.zfill(11)[:11] if len(d) <= 11 else d


def parse_fecha(v):
    """Julio trae celdas datetime de Excel. Acepta datetime/date/texto ISO."""
    if v is None or v == "":
        return None
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    s = str(v).strip().split()[0]
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        try:
            return dt.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            return None
    return None


# --- regimen_desc (texto) -> (tributo_codigo 4dig, regimen canonico) ---------
def regimen_a_codigo(desc):
    s = mayus(desc) or ""
    if "MYPE" in s:
        return "3121", "Régimen MYPE Tributario (RMT)"
    if "ESPECIAL" in s:
        return "3111", "Régimen Especial (RER)"
    if "GENERAL" in s:
        return "3031", "Régimen General"
    if "RUS" in s:
        return "4100", "RUS"
    if "AMAZON" in s:
        return "3311", "Amazonía"
    if "AGRAR" in s:
        return "3411", "Agrario"
    if "FRONTERA" in s:
        return "3611", "Frontera"
    return None, (texto(desc) or None)  # desconocido: sin codigo, conserva texto


# --- Resolver nombre -> ubigeo ----------------------------------------------
def norm(s):
    s = unicodedata.normalize("NFD", str(s or ""))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = s.upper().strip()
    s = re.sub(r"[.,]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


# Alias de abreviaturas/variantes frecuentes de SUNAT (distrito o provincia).
ALIAS = {
    "NASCA": "NAZCA",
    "CRL GREG ALBARRACIN LANCHIPA": "CORONEL GREGORIO ALBARRACIN LANCHIPA",
    "PROV CONST DEL CAL": "CALLAO",
    "PROV CONST DEL CALLAO": "CALLAO",
}


class Resolver:
    def __init__(self, catalogo):
        self.triple, self.provdist, self.distonly = {}, {}, {}
        for x in catalogo:
            u = str(x["u"])
            if not re.fullmatch(r"\d{6}", u):
                continue
            dep, prov, dist = norm(x["dep"]), norm(x["p"]), norm(x["d"])
            self.triple[(dep, prov, dist)] = u
            self.provdist.setdefault((prov, dist), set()).add(u)
            self.distonly.setdefault(dist, set()).add(u)

    def _variantes(self, name):
        """Nombre normalizado + variantes: alias, y con/sin parentesis."""
        n = norm(name)
        out = [n]
        if n in ALIAS:
            out.append(ALIAS[n])
        m = re.match(r"^(.*?)\s*\((.*?)\)\s*$", n)
        if m:
            fuera, dentro = m.group(1).strip(), m.group(2).strip()
            out += [fuera, dentro, ALIAS.get(fuera, fuera), ALIAS.get(dentro, dentro)]
        return [v for v in dict.fromkeys(out) if v]

    def resolver(self, dep, prov, dist):
        deps = self._variantes(dep)
        provs = self._variantes(prov)
        dists = self._variantes(dist)
        # 1) triple exacto (probando variantes)
        for de in deps:
            for pr in provs:
                for di in dists:
                    u = self.triple.get((de, pr, di))
                    if u:
                        return u, "triple"
        # 2) (provincia, distrito) unico
        for pr in provs:
            for di in dists:
                s = self.provdist.get((pr, di))
                if s and len(s) == 1:
                    return next(iter(s)), "prov+dist"
        # 3) distrito unico en el pais
        for di in dists:
            s = self.distonly.get(di)
            if s and len(s) == 1:
                return next(iter(s)), "dist-unico"
        return None, "no-resuelto"


# --- Lectura + mapeo del XLSX de julio --------------------------------------
IQ = {"160101": "Iquitos", "160108": "Punchana", "160112": "Belen",
      "160113": "San Juan Bautista"}


def domicilio_ubic(dom):
    """3 ultimos segmentos de 'domicilio' -> (distrito, provincia, departamento)."""
    segs = [s.strip() for s in str(dom or "").split(",") if s.strip()]
    if len(segs) < 3:
        return None, None, None
    return segs[-3], segs[-2], segs[-1]


def leer(archivo, tipo, resolver: Resolver, limit=None):
    wb = load_workbook(archivo, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    it = ws.iter_rows(values_only=True)
    hdr = [texto(h) for h in next(it)]
    idx = {h: i for i, h in enumerate(hdr) if h}

    def col(fila, nombre):
        i = idx.get(nombre)
        return fila[i] if (i is not None and i < len(fila)) else None

    regs = []
    metodo = Counter()
    sin_ubigeo = []            # (dep, prov, dist)
    regimen_desc_vals = Counter()
    iquitos = Counter()        # ubigeo IQ -> total
    iquitos_null = Counter()   # (dist embebido) cuando cae en Loreto/Maynas pero NO resolvio
    leidos = 0

    for fila in it:
        if fila is None or all(c is None for c in fila):
            continue
        ruc = ruc_norm(col(fila, "RUC"))
        if not ruc:
            continue
        leidos += 1

        if tipo == "juridica":
            razon = mayus(col(fila, "RAZON SOCIAL"))
            dom = texto(col(fila, "domicilio"))
            dist, prov, dep = domicilio_ubic(dom)
            trib_cod, regimen = regimen_a_codigo(col(fila, "regimen_desc"))
            regimen_desc_vals[mayus(col(fila, "regimen_desc")) or "(vacio)"] += 1
            direccion_texto = minus(dom)
        else:
            razon = mayus(col(fila, "APELLIDOS Y NOMBRES"))
            dep = mayus(col(fila, "Departamento"))
            prov = mayus(col(fila, "Provincia"))
            dist = mayus(col(fila, "Distrito"))
            trib_cod, regimen, direccion_texto = None, None, None

        ubigeo, met = resolver.resolver(dep, prov, dist)
        metodo[met] += 1
        if ubigeo is None:
            sin_ubigeo.append((dep, prov, dist))

        # Diagnostico Iquitos: por nombre embebido (Loreto/Maynas + 4 distritos)
        if norm(dep) == "LORETO" and norm(prov) == "MAYNAS" and \
                norm(dist) in {norm(v) for v in IQ.values()}:
            if ubigeo in IQ:
                iquitos[ubigeo] += 1
            else:
                iquitos_null[dist] += 1

        regs.append({
            "tipo": tipo, "ruc": ruc, "razon_social": razon,
            "ciiu": texto(col(fila, "CIIUr4_P")),
            "descripcion": mayus(col(fila, "CIIUr4_P_desc")),
            "comercio_exterior": texto(col(fila, "MarcaComercioExt_desc")),
            "ubigeo": ubigeo,
            "distrito": mayus(dist), "provincia": mayus(prov), "departamento": mayus(dep),
            "mes_inscripcion": PERIODO,
            "fecha_inscripcion": parse_fecha(col(fila, "fecInscOrig")),
            "fecha_inicio_actividades": parse_fecha(col(fila, "FecInicio")),
            "nombre_comercial": texto(col(fila, "NombreComercial")),
            "tipo_contribuyente": texto(col(fila, "TipContrib_desc")),
            "direccion_texto": direccion_texto,
            "tributo_codigo": trib_cod, "regimen": regimen,
        })
        if limit and leidos >= limit:
            break

    wb.close()
    return {"regs": regs, "leidos": leidos, "metodo": metodo,
            "sin_ubigeo": sin_ubigeo, "regimen_desc": regimen_desc_vals,
            "iquitos": iquitos, "iquitos_null": iquitos_null, "headers": hdr}


INSERT_COLS = ["tipo", "ruc", "razon_social", "ciiu", "descripcion",
               "comercio_exterior", "ubigeo", "distrito", "provincia", "departamento",
               "mes_inscripcion", "fecha_inscripcion", "fecha_inicio_actividades",
               "nombre_comercial", "tipo_contribuyente", "direccion_texto",
               "tributo_codigo", "regimen"]


async def insertar(database_url, regs):
    import asyncpg
    conn = await asyncpg.connect(database_url)
    try:
        antes = await conn.fetchval("SELECT COUNT(*) FROM nuevos_negocios")
        ph = ", ".join(f"${i}" for i in range(1, len(INSERT_COLS) + 1))
        sql = (f"INSERT INTO nuevos_negocios ({', '.join(INSERT_COLS)}) "
               f"VALUES ({ph}) ON CONFLICT (ruc) DO NOTHING")
        for i in range(0, len(regs), LOTE):
            lote = regs[i:i + LOTE]
            await conn.executemany(sql, [[r[c] for c in INSERT_COLS] for r in lote])
        despues = await conn.fetchval("SELECT COUNT(*) FROM nuevos_negocios")
        return antes, despues
    finally:
        await conn.close()


def reporte(res, tipo):
    print(f"  Filas leidas con RUC: {res['leidos']}")
    print(f"  Metodo de resolucion: {dict(res['metodo'])}")
    resueltos = res["leidos"] - res["metodo"].get("no-resuelto", 0)
    pct = 100 * resueltos / res["leidos"] if res["leidos"] else 0
    print(f"  ubigeo RESUELTO: {resueltos} ({pct:.2f}%) | NULL: {res['metodo'].get('no-resuelto',0)}")
    if res["sin_ubigeo"]:
        print("  Top no-resueltos (dep, prov, dist):")
        for k, n in Counter(res["sin_ubigeo"]).most_common(10):
            print(f"    {n:5}  {k}")
    if tipo == "juridica":
        print("  regimen_desc -> codigo:")
        for k, n in res["regimen_desc"].most_common():
            cod, _ = regimen_a_codigo(k)
            print(f"    {n:6}  {k!r:38} -> {cod}")
        print("  *** IQUITOS (Loreto/Maynas), 4 distritos ***")
        total_iq = sum(res["iquitos"].values()) + sum(res["iquitos_null"].values())
        for u, nom in IQ.items():
            print(f"    {u} {nom:20} resueltas={res['iquitos'].get(u,0)}")
        print(f"    NO resueltas (nombre Iquitos-Maynas pero ubigeo NULL): {dict(res['iquitos_null'])}")
        print(f"    TOTAL juridicas julio en los 4 distritos: {total_iq}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--archivo", required=True)
    ap.add_argument("--tipo", required=True, choices=["natural", "juridica"])
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--limit", type=int, default=None)
    ap.add_argument("--database-url", default=None)
    args = ap.parse_args()
    if not Path(args.archivo).exists():
        sys.exit(f"No existe: {args.archivo}")

    cargar_dotenv()
    catalogo = json.loads(DISTRITOS_JSON.read_text(encoding="utf-8"))
    resolver = Resolver(catalogo)
    print(f"Leyendo {args.archivo} (tipo={args.tipo})...")
    res = leer(args.archivo, args.tipo, resolver, args.limit)
    reporte(res, args.tipo)

    if args.dry_run:
        print("DRY-RUN: no se inserto nada.")
        m = res["regs"][0] if res["regs"] else {}
        print("  Primer registro mapeado:")
        for c in INSERT_COLS:
            print(f"    {c:22}= {m.get(c)}")
        return

    database_url = args.database_url or os.getenv("DATABASE_URL", "").strip()
    if not database_url:
        sys.exit("Falta DATABASE_URL.")
    print(f"Insertando {len(res['regs'])} filas por lotes de {LOTE} (ON CONFLICT ruc DO NOTHING)...")
    antes, despues = asyncio.run(insertar(database_url, res["regs"]))
    print(f"Listo. nuevos_negocios: {antes} -> {despues} (+{despues-antes} insertadas, "
          f"{len(res['regs'])-(despues-antes)} duplicadas ignoradas)")


if __name__ == "__main__":
    main()
